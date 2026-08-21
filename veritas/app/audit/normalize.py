"""Stage 1 — Ingest/Normalizer (architecture §7.1).

Deterministic data work: reads the stored raw upload bytes, parses the supported
formats (CSV/JSON/XLSX/Parquet) into a single consistent tabular view, and
writes it as a versioned JSON artifact. This is the ONLY stage that touches the
raw payload. No LLM is involved (llm_assist=False); the output is fully
reproducible so downstream stages can be re-run without re-ingesting.

Memory-light: rows are built in one pass and the whole normalized view is what
the matcher needs anyway; 100 MB cap bounds the working set.
"""
from __future__ import annotations
import csv
import json
from pathlib import Path

NORMALIZE_TEMPLATE_ID = "ingest-normalize-v1"
AGENT = "ingest_normalizer"


class NormalizeError(Exception):
    """Raised when the payload cannot be normalized into a tabular view."""


def normalize_csv(content: bytes) -> dict:
    """CSV → rows of dicts (first row = header)."""
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(text.splitlines(), strict=True)
    rows = [dict(r) for r in reader]
    return {"columns": list(reader.fieldnames or []), "rows": rows}


def normalize_json(content: bytes) -> dict:
    data = json.loads(content.decode("utf-8", errors="replace"))
    if isinstance(data, list):
        rows = [r for r in data if isinstance(r, dict)]
    elif isinstance(data, dict):
        # Tolerate a top-level envelope key; otherwise treat the object as one row.
        candidate = None
        for key in ("records", "rows", "ledger", "data"):
            if isinstance(data.get(key), list) and data.get(key):
                candidate = data.get(key)
                break
        if candidate is not None:
            rows = [r for r in candidate if isinstance(r, dict)]
        else:
            rows = [data]
    else:
        raise NormalizeError("JSON root must be an object or an array of objects")
    cols = sorted({c for r in rows for c in r.keys()}) if rows else []
    return {"columns": cols, "rows": rows}


def normalize_xlsx(content: bytes) -> dict:
    from openpyxl import load_workbook
    from io import BytesIO

    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            return {"columns": [], "rows": []}
        columns = [str(h) if h is not None else f"col{i}" for i, h in enumerate(header)]
        rows = [
            {columns[i]: row[i] for i in range(len(columns)) if i < len(row)}
            for row in it
        ]
        return {"columns": columns, "rows": rows}
    finally:
        wb.close()


def normalize_parquet(content: bytes) -> dict:
    import pyarrow.parquet as pq
    from io import BytesIO

    table = pq.read_table(BytesIO(content))
    columns = [c for c in table.column_names]
    rows = [dict(zip(columns, r)) for r in table.to_pylist()]
    return {"columns": columns, "rows": rows}


_FORMAT_PARSERS = {
    "csv": normalize_csv,
    "json": normalize_json,
    "xlsx": normalize_xlsx,
    "parquet": normalize_parquet,
}


def normalize(content: bytes, *, kind: str, source: str | None = None) -> dict:
    """Produce the normalized single-table view for a raw payload."""
    parser = _FORMAT_PARSERS.get(kind)
    if parser is None:
        raise NormalizeError(f"unsupported format: {kind!r}")
    parsed = parser(content)
    rows = parsed["rows"]
    result = {
        "table": "ledger",
        "columns": parsed["columns"],
        "rows": rows,
        "row_count": len(rows),
        "source": source or "upload",
        "format": kind,
        "schema_version": 1,
    }
    return result


def kind_from_path(filename: str) -> str:
    """Map an upload filename to the normalize kind (same allowlist as gate 1)."""
    suffix = Path(filename).suffix.lower().lstrip(".")
    if suffix not in _FORMAT_PARSERS:
        raise NormalizeError(f"unsupported upload type: {suffix!r}")
    return suffix

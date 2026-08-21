"""Safe read-only parse gate (architecture §6.2, third gate).

Run by the upload path INSIDE a sandboxed subprocess with hard memory and CPU
time limits (set via RLIMIT in ``service.py``). This module is the child entry
point: ``python -m app.uploads.safe_parse <kind> <path>`` — exits 0 if the file
parses cleanly under the safe options, non-zero with a message on stderr
otherwise.

Safety properties:
  * CSV  — stdlib ``csv`` reader, streaming, strict UTF-8.
  * JSON — stdlib ``json.load`` (memory bounded by the subprocess RLIMIT).
  * XLSX — openpyxl ``read_only=True, data_only=True``: no formula evaluation
           (only cached values are ever read), no external links, streaming.
  * Parquet — pyarrow ``ParquetFile`` with ``memory_map=False`` (no mmap of the
           whole file), reading only metadata + a bounded number of rows.

Nothing here shells out, evaluates formulas, or opens network connections. The
row caps keep the probe bounded — deep parsing happens in the pipeline's
normalize stage.
"""
from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

CSV = "csv"
JSON = "json"
XLSX = "xlsx"
PARQUET = "parquet"
KINDS = (CSV, JSON, XLSX, PARQUET)

DEFAULT_ROW_CAP = 100_000
DEFAULT_SHEET_CAP = 64


def _probe_csv(path: Path, row_cap: int) -> None:
    # Bound the per-field size: a single pathological field (e.g. one enormous
    # column) is quarantined instead of being buffered into memory.
    csv.field_size_limit(1 << 20)  # 1 MiB per field
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        for i, row in enumerate(reader):
            if i >= row_cap:
                break
            # Force materialization so lazy parsing still validates the row.
            _ = list(row)
    # A malformed CSV (e.g. unclosed quote or oversized field) raises here
    # via csv.Error; invalid UTF-8 raises UnicodeDecodeError mid-iteration.


def _probe_json(path: Path) -> None:
    with path.open("r", encoding="utf-8") as fh:
        json.load(fh)


def _probe_xlsx(path: Path, row_cap: int, sheet_cap: int) -> None:
    from openpyxl import load_workbook

    # read_only=True: streaming parser, no formula evaluation; data_only=True:
    # cached cell values only, never computed. No external links are loaded.
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_idx, ws in enumerate(wb.worksheets):
            if sheet_idx >= sheet_cap:
                break
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= row_cap:
                    break
                _ = list(row)
    finally:
        wb.close()


def _probe_parquet(path: Path, row_cap: int) -> None:
    import pyarrow.parquet as pq

    pf = pq.ParquetFile(path, memory_map=False)
    _ = pf.schema  # forces footer parse — proves this is really a parquet file
    rows = 0
    for batch in pf.iter_batches(batch_size=1000):
        rows += batch.num_rows
        if rows >= row_cap:
            break


def probe(kind: str, path: Path, row_cap: int = DEFAULT_ROW_CAP) -> None:
    """Parse ``path`` per ``kind``; raise ValueError on any failure."""
    if kind == CSV:
        _probe_csv(path, row_cap)
    elif kind == JSON:
        _probe_json(path)
    elif kind == XLSX:
        _probe_xlsx(path, row_cap, DEFAULT_SHEET_CAP)
    elif kind == PARQUET:
        _probe_parquet(path, row_cap)
    else:
        raise ValueError(f"unknown parse kind: {kind!r}")


def main(argv: list[str] | None = None) -> int:
    args = argv if argv is not None else sys.argv[1:]
    if len(args) != 2:
        print("usage: python -m app.safe_parse <kind> <path>", file=sys.stderr)
        return 2
    kind, raw_path = args
    try:
        probe(kind, Path(raw_path))
    except Exception as exc:  # any failure means: do not pass downstream
        print(f"parse failed for {kind}: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
